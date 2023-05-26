#In this tutorial, we will be focusing on how to implement CRUD (create read update delete) operation using python and excel as database
#for this, we will be making a task reminder project.
#Getting started, please create a folder with 3 python files for code and 1 png file to use as icon in the notification

#Here, we will code to create a value and save it in the excel sheet
from openpyxl import Workbook
from openpyxl import load_workbook
import os

#function to check if the excel file already exists or not
def check_exists():
    return os.path.exists('tasks.xlsx')

#function to return the first non populated row of the sheet
def get_first_nonpopulated(sheet):
    column = sheet['A'] #Assuming that column A is always populated
    for cell in column[::-1]:
        if cell.value is None:
            return cell.row
    return 1

def main_program():
    check_exist = check_exists()
    if check_exist:
        workbook = load_workbook('tasks.xlsx')
        sheet = workbook.active
        
        #Remember to add 2 because the function returns the index of the last populated row and we need to add the values in the new cells
        first_nonpopulated = get_first_nonpopulated(sheet) + 2
        
        task = input("Enter the task: ")
        date = input("Enter the date (YYYY-MM-DD): ")
        time = input("Enter the time (HH:MM): ")
        
        sheet.cell(row = first_nonpopulated, column = 1, value = task)
        sheet.cell(row = first_nonpopulated, column = 2, value = date)
        sheet.cell(row = first_nonpopulated, column = 3, value = time)
        
        workbook.save('tasks.xlsx')
        
    else:
        workbook = Workbook()
        sheet = workbook.active
            
        #if the file doesn't exists, we need to enter the values Task, Date, Time in cells A1, B1, C1 respectively.
        sheet['A1'] = "Task"
        sheet['B1'] = "Date"
        sheet['C1'] = "Time"
        
        task = input("Enter the task: ")
        date = input("Enter the date (YYYY-MM-DD): ")
        time = input("Enter the time (HH:MM): ")
            
        row = [task, date, time]
            
        sheet.append(row)
        workbook.save('tasks.xlsx')
            
main_program()

#with this, we just created a code to implement the create operation as it creates (inputs) a value to store in the excel            
            
        
        