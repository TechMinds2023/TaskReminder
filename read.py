#Here, we will write the code to read the values from the file and notify if the date and time matches

from datetime import datetime
from openpyxl import load_workbook
from win10toast import ToastNotifier
import time

#function to read the values
def read():
    workbook = load_workbook('tasks.xlsx')
    sheet = workbook.active
    
    #get the current date and time and convert it to string
    current_date = datetime.now().date().strftime('%Y-%m-%d')
    current_time = datetime.now().time().strftime('%H:%M')
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        task = row[0]
        date = row[1]
        time = row[2]
        
        if date == current_date and time == current_time:
            toaster = ToastNotifier()
            toaster.show_toast("Task Reminder", f"it's time to complete your task: {task}", 'python_icon.ico', duration = 10)
            #Notification may not appear in the recording, so just printing a statement to confirm that the code is working
            print("Notified!!")

while True:
    read()
    #Program should recheck only after 1 minute
    time.sleep(60)
    
    #with this, we come to an end of this video, now let's test the code by running it
    #we have to wait for a minute as the current time is 13:54 and we entered 13:55
    #So as you can see, the program has notified us exactly on the time we entered in the sheet, you may ignore the warning, because it only appears in some IDEs
    #So that is it for today, we will be covering update and delete operation in the next video, thanks for watching!!!